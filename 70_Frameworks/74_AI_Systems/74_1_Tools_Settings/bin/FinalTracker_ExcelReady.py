import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import json
import os
import subprocess
import platform
from datetime import datetime, timedelta
from collections import defaultdict
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from matplotlib import font_manager
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from tkcalendar import DateEntry
    HAS_CALENDAR = True
except ImportError:
    HAS_CALENDAR = False

# 日本語フォント設定
plt.rcParams['font.sans-serif'] = ['Yu Gothic', 'MS Gothic', 'Hiragino Sans', 'IPAexGothic']
plt.rcParams['axes.unicode_minus'] = False

# ==========================================
# strat-lab システム専用パス固定定義
# ==========================================
SETTINGS_FILE = r"C:\Users\akasaka.kazuyuki\OneDrive - ユーザーサイド株式会社\strat-lab\strat-lab\70_Frameworks\74_AI_Systems\74_1_Tools_Settings\configs/tracker_settings.json"
DEFAULT_LOG_FILE = 'work_log.xlsx'
FIXED_OUTPUT_DIR = r"C:\Users\akasaka.kazuyuki\OneDrive - ユーザーサイド株式会社\strat-lab\strat-lab\10_Daily\11_工数管理\Pythonログ"

class ModernTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("工数管理 Pro")
        self.root.geometry("420x450")
        self.root.minsize(400, 320)
        self.root.attributes("-topmost", True)
        self.root.configure(bg="#1a1a2e")

        self.current_task = tk.StringVar(value="停止中")
        self.current_memo = ""
        self.start_time = None
        self.groups = {}
        self.expanded_groups = {}
        self.output_dir = FIXED_OUTPUT_DIR

        self.load_settings()
        self.setup_ui()

    def load_settings(self):
        os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)
        
        default = {
            "groups": {
                "メイン業務": ["電話対応", "事務処理", "会議", "資料作成"],
                "その他": ["休憩", "移動", "雑務"]
            },
            "output_dir": self.output_dir
        }
        
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf_8') as f:
                settings = json.load(f)
                if "groups" in settings:
                    self.groups = settings["groups"]
                    self.output_dir = FIXED_OUTPUT_DIR
                else:
                    self.groups = settings
                    self.output_dir = FIXED_OUTPUT_DIR
        else:
            self.groups = default["groups"]
            self.output_dir = FIXED_OUTPUT_DIR
            self.save_settings()
        
        for group in self.groups.keys():
            self.expanded_groups[group] = True

    def save_settings(self):
        settings = {
            "groups": self.groups,
            "output_dir": self.output_dir
        }
        with open(SETTINGS_FILE, 'w', encoding='utf_8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)
    
    def get_log_file_path(self):
        return os.path.join(self.output_dir, DEFAULT_LOG_FILE)
    
    def parse_date_safe(self, date_str):
        if not date_str or not str(date_str).strip(): return None
        date_str = str(date_str).strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
            try: return datetime.strptime(date_str.split()[0], fmt.split()[0]).date()
            except: continue
        return None

    def setup_ui(self):
        status_frame = tk.Frame(self.root, bg="#16213e", bd=0, relief="flat")
        status_frame.pack(fill="x", padx=0, pady=0)
        
        status_inner = tk.Frame(status_frame, bg="#16213e")
        status_inner.pack(fill="x", padx=10, pady=6)
        
        status_label = tk.Label(status_inner, text="現在:", bg="#16213e", fg="#94a3b8", font=("Yu Gothic", 9))
        status_label.pack(side="left", padx=(0, 6))
        
        task_label = tk.Label(status_inner, textvariable=self.current_task, font=("Yu Gothic", 10, "bold"), bg="#16213e", fg="#38bdf8")
        task_label.pack(side="left")
        
        memo_btn = tk.Button(status_inner, text="📝", width=3, command=self.add_memo, bg="#0f172a", fg="#f1f5f9", activebackground="#1e293b", relief="flat", font=("Yu Gothic", 10), cursor="hand2", borderwidth=0, highlightthickness=0, pady=2)
        memo_btn.pack(side="right", padx=2)

        bottom_frame = tk.Frame(self.root, bg="#1a1a2e")
        bottom_frame.pack(side="bottom", fill="x", padx=8, pady=5)
        
        btn_row = tk.Frame(bottom_frame, bg="#1a1a2e")
        btn_row.pack(fill="x")
        
        complete_btn = tk.Button(btn_row, text="✓ 完了", command=self.complete_day, bg="#3b82f6", fg="white", font=("Yu Gothic", 9, "bold"), relief="flat", bd=0, activebackground="#2563eb", cursor="hand2", padx=8, pady=5, borderwidth=0, highlightthickness=0)
        complete_btn.pack(side="left", padx=(0, 3))
        
        analyze_btn = tk.Button(btn_row, text="📊", command=self.open_analysis, bg="#8b5cf6", fg="white", font=("Yu Gothic", 11), relief="flat", bd=0, width=3, activebackground="#7c3aed", cursor="hand2", pady=5, borderwidth=0, highlightthickness=0)
        analyze_btn.pack(side="left", padx=(0, 3))
        
        open_folder_btn = tk.Button(btn_row, text="📁", command=self.open_output_folder, bg="#f59e0b", fg="white", font=("Yu Gothic", 11), relief="flat", bd=0, width=3, activebackground="#d97706", cursor="hand2", pady=5, borderwidth=0, highlightthickness=0)
        open_folder_btn.pack(side="left", padx=(0, 2))
        
        change_folder_btn = tk.Button(btn_row, text="⚙", width=3, command=self.change_output_folder, bg="#f59e0b", fg="white", font=("Yu Gothic", 10), relief="flat", bd=0, activebackground="#d97706", cursor="hand2", pady=5, borderwidth=0, highlightthickness=0)
        change_folder_btn.pack(side="left", padx=(0, 3))
        
        add_group_btn = tk.Button(btn_row, text="+ グループ", command=self.add_group, bg="#10b981", fg="white", font=("Yu Gothic", 9, "bold"), relief="flat", bd=0, activebackground="#059669", cursor="hand2", padx=8, pady=5, borderwidth=0, highlightthickness=0)
        add_group_btn.pack(side="left")

        scroll_container = tk.Frame(self.root, bg="#1a1a2e")
        scroll_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        self.canvas = tk.Canvas(scroll_container, bg="#1a1a2e", highlightthickness=0, bd=0)
        self.scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = tk.Frame(self.canvas, bg="#1a1a2e")
        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=8)
        self.scrollbar.pack(side="right", fill="y")
        self.refresh_buttons()

    def refresh_buttons(self):
        for widget in self.scroll_frame.winfo_children(): widget.destroy()
        group_list = list(self.groups.keys())
        for group_idx, group in enumerate(group_list):
            tasks = self.groups[group]
            header_frame = tk.Frame(self.scroll_frame, bg="#0f172a", bd=0, relief="flat")
            header_frame.pack(fill="x", padx=4, pady=6)
            header_inner = tk.Frame(header_frame, bg="#0f172a")
            header_inner.pack(fill="x", padx=10, pady=8)
            
            arrow_frame = tk.Frame(header_inner, bg="#0f172a")
            arrow_frame.pack(side="left", padx=(0, 8))
            if group_idx > 0:
                tk.Button(arrow_frame, text="▲", width=1, command=lambda g=group: self.move_group_up(g), bg="#1e293b", fg="#94a3b8", relief="flat", font=("Yu Gothic", 7), borderwidth=0).pack(side="top", pady=(0, 1))
            if group_idx < len(group_list) - 1:
                tk.Button(arrow_frame, text="▼", width=1, command=lambda g=group: self.move_group_down(g), bg="#1e293b", fg="#94a3b8", relief="flat", font=("Yu Gothic", 7), borderwidth=0).pack(side="top")

            toggle_symbol = "▼" if self.expanded_groups.get(group, True) else "▶"
            tk.Button(header_inner, text=toggle_symbol, width=2, command=lambda g=group: self.toggle_group(g), bg="#0f172a", fg="#cbd5e1", relief="flat", font=("Yu Gothic", 9), borderwidth=0).pack(side="left", padx=(0, 8))
            tk.Label(header_inner, text=group, bg="#0f172a", fg="#f1f5f9", font=("Yu Gothic", 10, "bold")).pack(side="left")

            btn_frame = tk.Frame(header_inner, bg="#0f172a")
            btn_frame.pack(side="right")
            tk.Button(btn_frame, text="✎", width=2, command=lambda g=group: self.edit_group_name(g), bg="#0f172a", fg="#64748b", relief="flat", font=("Yu Gothic", 10), borderwidth=0).pack(side="left", padx=2)
            tk.Button(btn_frame, text="×", width=2, command=lambda g=group: self.delete_group(g), bg="#0f172a", fg="#ef4444", relief="flat", font=("Yu Gothic", 11), borderwidth=0).pack(side="left", padx=2)
            tk.Button(btn_frame, text="+", width=2, command=lambda g=group: self.add_task(g), bg="#0f172a", fg="#10b981", relief="flat", font=("Yu Gothic", 11), borderwidth=0).pack(side="left", padx=2)

            if self.expanded_groups.get(group, True):
                task_container = tk.Frame(self.scroll_frame, bg="#1a1a2e")
                task_container.pack(fill="x", padx=4, pady=(0, 4))
                for task_idx, task in enumerate(tasks):
                    row, col = task_idx // 2, task_idx % 2
                    task_cell = tk.Frame(task_container, bg="#1a1a2e")
                    task_cell.grid(row=row, column=col, padx=3, pady=2, sticky="ew")
                    task_container.grid_columnconfigure(col, weight=1)
                    
                    task_arrow_frame = tk.Frame(task_cell, bg="#1a1a2e")
                    task_arrow_frame.pack(side="left", padx=(2, 4))
                    if task_idx > 0:
                        tk.Button(task_arrow_frame, text="▲", width=1, command=lambda g=group, t=task: self.move_task_up(g, t), bg="#1e293b", fg="#64748b", relief="flat", font=("Yu Gothic", 6), borderwidth=0).pack(side="top", pady=(0, 1))
                    if task_idx < len(tasks) - 1:
                        tk.Button(task_arrow_frame, text="▼", width=1, command=lambda g=group, t=task: self.move_task_down(g, t), bg="#1e293b", fg="#64748b", relief="flat", font=("Yu Gothic", 6), borderwidth=0).pack(side="top")

                    tk.Button(task_cell, text=task, anchor="w", command=lambda t=task: self.switch_task(t), bg="#16213e", fg="#e2e8f0", relief="flat", font=("Yu Gothic", 9), padx=8, pady=7, borderwidth=0).pack(side="left", fill="both", expand=True)
                    action_f = tk.Frame(task_cell, bg="#16213e")
                    action_f.pack(side="left", padx=2)
                    tk.Button(action_f, text="✎", width=2, command=lambda g=group, t=task: self.edit_task(g, t), bg="#16213e", fg="#64748b", relief="flat", borderwidth=0).pack(side="left")
                    tk.Button(action_f, text="×", width=2, command=lambda g=group, t=task: self.delete_task(g, t), bg="#16213e", fg="#ef4444", relief="flat", borderwidth=0).pack(side="left")

    def move_group_up(self, group):
        gl = list(self.groups.keys()); idx = gl.index(group)
        if idx > 0:
            gl[idx], gl[idx-1] = gl[idx-1], gl[idx]
            self.groups = {g: self.groups[g] for g in gl}; self.save_settings(); self.refresh_buttons()

    def move_group_down(self, group):
        gl = list(self.groups.keys()); idx = gl.index(group)
        if idx < len(gl)-1:
            gl[idx], gl[idx+1] = gl[idx+1], gl[idx]
            self.groups = {g: self.groups[g] for g in gl}; self.save_settings(); self.refresh_buttons()

    def move_task_up(self, group, task):
        idx = self.groups[group].index(task)
        if idx > 0:
            self.groups[group][idx], self.groups[group][idx-1] = self.groups[group][idx-1], self.groups[group][idx]
            self.save_settings(); self.refresh_buttons()

    def move_task_down(self, group, task):
        idx = self.groups[group].index(task)
        if idx < len(self.groups[group])-1:
            self.groups[group][idx], self.groups[group][idx+1] = self.groups[group][idx+1], self.groups[group][idx]
            self.save_settings(); self.refresh_buttons()

    def toggle_group(self, group):
        self.expanded_groups[group] = not self.expanded_groups.get(group, True); self.refresh_buttons()

    def add_group(self):
        name = simpledialog.askstring("追加", "グループ名:"); 
        if name and name not in self.groups: self.groups[name] = []; self.expanded_groups[name] = True; self.save_settings(); self.refresh_buttons()

    def edit_group_name(self, old):
        new = simpledialog.askstring("編集", "新グループ名:", initialvalue=old)
        if new and new != old:
            new_g = {}; [new_g.update({(new if k==old else k): v}) for k,v in self.groups.items()]
            self.groups = new_g; self.expanded_groups[new] = self.expanded_groups.pop(old, True); self.save_settings(); self.refresh_buttons()

    def delete_group(self, group):
        if messagebox.askyesno("確認", f"「{group}」を削除？"): del self.groups[group]; self.save_settings(); self.refresh_buttons()

    def add_task(self, group):
        name = simpledialog.askstring("追加", f"「{group}」の新タスク:")
        if name and name not in self.groups[group]: self.groups[group].append(name); self.save_settings(); self.refresh_buttons()

    def edit_task(self, group, old):
        new = simpledialog.askstring("編集", "新タスク名:", initialvalue=old)
        if new and new != old: idx = self.groups[group].index(old); self.groups[group][idx] = new; self.save_settings(); self.refresh_buttons()

    def delete_task(self, group, task):
        if messagebox.askyesno("確認", f"「{task}」を削除？"): self.groups[group].remove(task); self.save_settings(); self.refresh_buttons()

    def add_memo(self):
        if self.current_task.get() == "停止中": return
        m = simpledialog.askstring("メモ", f"「{self.current_task.get()}」のメモ:", initialvalue=self.current_memo)
        if m is not None: self.current_memo = m

    def switch_task(self, name):
        now = datetime.now()
        if self.current_task.get() != "停止中": self.save_log(self.current_task.get(), self.start_time, now, self.current_memo)
        self.current_task.set(name); self.current_memo = ""; self.start_time = now

    def complete_day(self):
        if self.current_task.get() != "停止中":
            if messagebox.askyesno("確認", "完了しますか？"):
                self.save_log(self.current_task.get(), self.start_time, datetime.now(), self.current_memo)
                self.current_task.set("停止中"); self.current_memo = ""; self.start_time = None
                messagebox.showinfo("完了", f"保存先:\n{self.get_log_file_path()}")

    def open_output_folder(self):
        if os.path.exists(self.output_dir): os.startfile(self.output_dir)

    def change_output_folder(self):
        new = filedialog.askdirectory(title="保存先選択", initialdir=self.output_dir)
        if new: self.output_dir = new; self.save_settings(); messagebox.showinfo("変更", f"保存先:\n{new}")

    def save_log(self, task, start, end, memo=""):
        dur = round((end - start).total_seconds() / 60, 1)
        path = self.get_log_file_path()
        
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            
            if ws.tables:
                for table_name in list(ws.tables.keys()):
                    del ws.tables[table_name]
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["日付", "開始", "終了", "タスク", "分", "メモ"])
        
        next_row = ws.max_row + 1
        ws.append([
            start.strftime("%Y/%m/%d"),
            start.strftime("%H:%M"),
            end.strftime("%H:%M"),
            task,
            dur,
            memo
        ])
        
        table_range = f"A1:F{ws.max_row}"
        table = Table(displayName="WorkLog", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        wb.save(path)

    def open_analysis(self):
        log_path = self.get_log_file_path()
        if not os.path.exists(log_path):
            messagebox.showwarning("データなし", "ログファイルが見つかりません")
            return
        
        analysis_win = tk.Toplevel(self.root)
        analysis_win.title("工数分析")
        analysis_win.geometry("1000x700")
        analysis_win.configure(bg="#1a1a2e")
        
        header = tk.Label(analysis_win, text="📊 工数分析", font=("Yu Gothic", 16, "bold"), bg="#16213e", fg="#f1f5f9", pady=15)
        header.pack(fill="x")
        
        control_frame = tk.Frame(analysis_win, bg="#1a1a2e")
        control_frame.pack(fill="x", padx=20, pady=10)
        
        # 表示モード選択（初期値を"range"に変更）
        mode_var = tk.StringVar(value="range")
        
        # カレンダーを常に有効にして、表示/非表示で制御
        tk.Label(control_frame, text="表示:", bg="#1a1a2e", fg="#94a3b8", font=("Yu Gothic", 10)).pack(side="left", padx=(0, 5))
        tk.Radiobutton(control_frame, text="日別", variable=mode_var, value="daily", bg="#1a1a2e", fg="#f1f5f9", selectcolor="#16213e", font=("Yu Gothic", 10), activebackground="#1a1a2e", activeforeground="#38bdf8").pack(side="left", padx=5)
        tk.Radiobutton(control_frame, text="期間指定", variable=mode_var, value="range", bg="#1a1a2e", fg="#f1f5f9", selectcolor="#16213e", font=("Yu Gothic", 10), activebackground="#1a1a2e", activeforeground="#38bdf8").pack(side="left", padx=5)
        tk.Radiobutton(control_frame, text="全期間", variable=mode_var, value="all", bg="#1a1a2e", fg="#f1f5f9", selectcolor="#16213e", font=("Yu Gothic", 10), activebackground="#1a1a2e", activeforeground="#38bdf8").pack(side="left", padx=5)
        
        start_cal = None
        end_cal = None
        
        if HAS_CALENDAR:
            tk.Label(control_frame, text="開始:", bg="#1a1a2e", fg="#94a3b8", font=("Yu Gothic", 10)).pack(side="left", padx=(15, 5))
            start_cal = DateEntry(control_frame, width=12, background='#3b82f6', foreground='white', borderwidth=2, date_pattern='y/mm/dd', locale='ja_JP')
            start_cal.pack(side="left", padx=5)
            tk.Label(control_frame, text="終了:", bg="#1a1a2e", fg="#94a3b8", font=("Yu Gothic", 10)).pack(side="left", padx=(10, 5))
            end_cal = DateEntry(control_frame, width=12, background='#3b82f6', foreground='white', borderwidth=2, date_pattern='y/mm/dd', locale='ja_JP')
            end_cal.pack(side="left", padx=5)
        
        def refresh_analysis():
            for w in chart_frame.winfo_children(): w.destroy()
            
            mode = mode_var.get()
            
            wb = load_workbook(log_path)
            ws = wb.active
            data = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    row_date = self.parse_date_safe(row[0])
                    if row_date:
                        data.append({
                            "日付": row_date,
                            "タスク": row[3],
                            "分": float(row[4]) if row[4] else 0
                        })
            
            if not data:
                tk.Label(chart_frame, text="データがありません", bg="#1a1a2e", fg="#94a3b8", font=("Yu Gothic", 12)).pack(expand=True)
                return
            
            # モードに応じてデータをフィルタ
            if mode == "daily":
                if HAS_CALENDAR and start_cal:
                    target_date = start_cal.get_date()
                    filtered_data = [d for d in data if d["日付"] == target_date]
                    title_suffix = f"({target_date})"
                else:
                    latest_date = max([d["日付"] for d in data])
                    filtered_data = [d for d in data if d["日付"] == latest_date]
                    title_suffix = f"({latest_date})"
            elif mode == "range":
                if HAS_CALENDAR and start_cal and end_cal:
                    start_date = start_cal.get_date()
                    end_date = end_cal.get_date()
                    filtered_data = [d for d in data if start_date <= d["日付"] <= end_date]
                    title_suffix = f"({start_date} 〜 {end_date})"
                else:
                    filtered_data = data
                    title_suffix = "(全期間)"
            else:  # all
                filtered_data = data
                title_suffix = "(全期間)"
            
            if not filtered_data:
                tk.Label(chart_frame, text="指定期間にデータがありません", bg="#1a1a2e", fg="#94a3b8", font=("Yu Gothic", 12)).pack(expand=True)
                return
            
            # タスク別集計
            task_time = defaultdict(float)
            for row in filtered_data:
                task_time[row["タスク"]] += row["分"]
            
            # カラーパレット（視認性の高い色）
            distinct_colors = [
                '#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8',
                '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788',
                '#E07A5F', '#81B29A', '#F2CC8F', '#A8DADC', '#E63946'
            ]
            
            # 2つのグラフ表示
            fig = plt.Figure(figsize=(12, 5), facecolor='#1a1a2e')
            
            # 1. タスク別工数（横棒グラフ）
            ax1 = fig.add_subplot(121, facecolor='#16213e')
            tasks = list(task_time.keys())
            times = [task_time[t]/60 for t in tasks]
            colors = [distinct_colors[i % len(distinct_colors)] for i in range(len(tasks))]
            ax1.barh(tasks, times, color=colors)
            ax1.set_xlabel('時間 (h)', color='#f1f5f9', fontsize=11)
            ax1.set_title(f'タスク別工数 {title_suffix}', color='#f1f5f9', fontweight='bold', fontsize=13)
            ax1.tick_params(colors='#f1f5f9', labelsize=10)
            ax1.spines['bottom'].set_color('#94a3b8')
            ax1.spines['left'].set_color('#94a3b8')
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.grid(axis='x', color='#2d3748', linestyle='--', linewidth=0.5, alpha=0.7)
            
            # 2. タスク別工数（円グラフ）
            ax2 = fig.add_subplot(122, facecolor='#16213e')
            sizes = [task_time[t]/60 for t in tasks]
            
            def autopct_format(pct):
                return f'{pct:.1f}%' if pct > 3 else ''
            
            wedges, texts, autotexts = ax2.pie(
                sizes, 
                labels=tasks, 
                autopct=autopct_format,
                colors=colors,
                textprops={'color': '#ffffff', 'fontsize': 10, 'weight': 'bold'},
                startangle=90,
                pctdistance=0.85
            )
            
            for autotext in autotexts:
                autotext.set_color('#000000')
                autotext.set_fontsize(11)
                autotext.set_weight('bold')
            
            for text in texts:
                text.set_fontsize(9)
            
            ax2.set_title(f'タスク割合 {title_suffix}', color='#f1f5f9', fontweight='bold', fontsize=13)
            
            fig.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Button(control_frame, text="更新", command=refresh_analysis, bg="#3b82f6", fg="white", font=("Yu Gothic", 10, "bold"), relief="flat", padx=15, pady=5, cursor="hand2", borderwidth=0).pack(side="left", padx=10)
        
        chart_frame = tk.Frame(analysis_win, bg="#1a1a2e")
        chart_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        refresh_analysis()

if __name__ == "__main__":
    root = tk.Tk()
    app = ModernTracker(root)
    root.mainloop()