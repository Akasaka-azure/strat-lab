# mask_personal_info.py

import re
import os
import sys
import csv
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import spacy
    nlp = spacy.load("ja_ginza")
    NLP_AVAILABLE = True
except Exception:
    NLP_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ==================== 設定 ====================

MASK_VALUE = "***"
OUTPUT_PREFIX = "【マスク済み】"
MAX_TEXT_LENGTH = 255

HEADER_KEYWORDS = [
    "名前", "氏名", "お名前", "姓", "名", "苗字",
    "ふりがな", "フリガナ", "よみがな", "ヨミガナ", "よみ", "かな", "カナ",
    "会社", "企業", "法人", "組織",
    "住所", "都道府県", "市区町村", "番地", "建物",
    "パスワード", "pw", "pass", "password",
    "電話", "tel", "phone", "携帯", "mobile", "phs",
    "fax", "ファックス", "ファクス",
    "mail", "メール", "email",
    "登録者", "担当者", "備考", "メモ", "郵便番号", "顧客コード", "会社コード",
]

TRUNCATE_KEYWORDS = ["メモ", "備考"]

PATTERNS = [
    re.compile(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'),
    re.compile(r'(\(?\d{2,5}\)?[-\s]?\d{1,4}[-\s]?\d{3,4})'),
    re.compile(r'0[5789]0[-\s]?\d{4}[-\s]?\d{4}'),
    re.compile(r'〒?\d{3}[-‐－]\d{4}'),
    re.compile(r'(北海道|東京都|(?:大阪|京都)府|.{2,3}県).{2,50}(丁目|番地|号|[-\d]+F)'),
]

NER_TARGET_LABELS = {"Person", "GPE", "Location", "Organization", "Facility"}

# ==================== マスク処理 ====================

def should_mask_header(header):
    if header is None:
        return False
    h = str(header).lower().strip()
    return any(kw.lower() in h for kw in HEADER_KEYWORDS)


def should_truncate_header(header):
    if header is None:
        return False
    h = str(header).strip()
    return any(kw in h for kw in TRUNCATE_KEYWORDS)


def mask_value(value):
    if value is None or str(value).strip() == "":
        return value
    return MASK_VALUE


def mask_by_pattern(value):
    if value is None:
        return value
    text = str(value)
    for pattern in PATTERNS:
        text = pattern.sub(MASK_VALUE, text)
    return text


def mask_by_ner(value):
    if not NLP_AVAILABLE or value is None:
        return mask_by_pattern(value)
    text = str(value)
    if not text.strip():
        return value
    doc = nlp(text)
    masked = text
    entities = sorted(doc.ents, key=lambda e: e.start_char, reverse=True)
    for ent in entities:
        if ent.label_ in NER_TARGET_LABELS:
            masked = masked[:ent.start_char] + MASK_VALUE + masked[ent.end_char:]
    return mask_by_pattern(masked)


def truncate(value):
    if value is None:
        return value
    text = str(value)
    return text[:MAX_TEXT_LENGTH] if len(text) > MAX_TEXT_LENGTH else text


# ==================== Excel処理 ====================

def process_xlsx(src_path, dst_path):
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxlがインストールされていません")

    wb = openpyxl.load_workbook(src_path)
    for ws in wb.worksheets:
        masked_cols = set()
        truncate_cols = set()
        for col in range(1, (ws.max_column or 0) + 1):
            header_val = ws.cell(row=1, column=col).value
            if should_mask_header(header_val):
                masked_cols.add(col)
            if should_truncate_header(header_val):
                truncate_cols.add(col)

        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue
                if cell.column in masked_cols:
                    cell.value = mask_value(cell.value)
                else:
                    cell.value = mask_by_ner(cell.value)
                if cell.column in truncate_cols and cell.value:
                    cell.value = truncate(cell.value)
    wb.save(dst_path)


# ==================== CSV処理 ====================

def process_csv(src_path, dst_path):
    encodings = ["utf-8-sig", "cp932", "shift_jis", "utf-8"]
    rows = None
    used_enc = "utf-8-sig"
    for enc in encodings:
        try:
            with open(src_path, newline="", encoding=enc) as f:
                rows = list(csv.reader(f))
            used_enc = enc
            break
        except Exception:
            continue

    if not rows:
        raise ValueError("CSVの読み込みに失敗しました")

    header = rows[0]
    masked_cols = {i for i, h in enumerate(header) if should_mask_header(h)}
    truncate_cols = {i for i, h in enumerate(header) if should_truncate_header(h)}

    result = [header]
    total = len(rows) - 1
    for idx, row in enumerate(rows[1:], 1):
        if idx % 50 == 0:
            print(f"  処理中... {idx}/{total}行")
        new_row = []
        for i, val in enumerate(row):
            if i in masked_cols:
                v = MASK_VALUE if val.strip() else val
            else:
                v = mask_by_ner(val)
            if i in truncate_cols:
                v = truncate(v)
            new_row.append(v)
        result.append(new_row)

    with open(dst_path, "w", newline="", encoding=used_enc) as f:
        csv.writer(f).writerows(result)


# ==================== メイン ====================

def main(file_path=None):
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        file_path = filedialog.askopenfilename(
            title="マスク対象ファイルを選択",
            filetypes=[("対応ファイル", "*.xlsx *.xls *.csv"), ("すべて", "*.*")]
        )
        root.destroy()

    if not file_path:
        return

    if not os.path.exists(file_path):
        messagebox.showerror("エラー", f"ファイルが見つかりません:\n{file_path}")
        return

    if not NLP_AVAILABLE:
        answer = messagebox.askyesno(
            "GiNZA未インストール",
            "GiNZA（高精度NLPマスク）が使えません。\n正規表現のみの簡易マスクで続行しますか？"
        )
        if not answer:
            return

    dir_name = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    dst_path = os.path.join(dir_name, OUTPUT_PREFIX + base_name)
    ext = os.path.splitext(file_path)[1].lower()

    print(f"📂 処理対象: {file_path}")
    print(f"📝 出力先  : {dst_path}")
    print(f"🧠 NLPモード: {'GiNZA有効' if NLP_AVAILABLE else '正規表現のみ'}")

    try:
        if ext in [".xlsx", ".xls"]:
            if ext == ".xls":
                tmp = dst_path.replace(".xls", ".xlsx")
                shutil.copy2(file_path, tmp)
                process_xlsx(tmp, tmp)
                dst_path = tmp
            else:
                process_xlsx(file_path, dst_path)
        elif ext == ".csv":
            process_csv(file_path, dst_path)
        else:
            messagebox.showerror("エラー", f"非対応の形式です: {ext}")
            return

        messagebox.showinfo("完了", f"マスク済みファイルを保存しました:\n{dst_path}")
        print("🎉 完了")
        sys.exit(0)

    except Exception as e:
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
        print(f"❌ エラー: {e}")
        sys.exit(1)


if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else None
    main(file_path)